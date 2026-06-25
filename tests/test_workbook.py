from openpyxl import load_workbook

from reporting.workbook import write_workbook
from rules.engine import run_all


def test_workbook_structure(ctx, registry, tmp_path):
    findings = run_all(ctx)
    path = write_workbook(findings, registry, tmp_path / "exceptions.xlsx",
                          run_label="test run")
    wb = load_workbook(path)
    assert wb.sheetnames == ["Summary", "CRITICAL", "HIGH", "MEDIUM", "INFO",
                             "Methodology", "Run Info"]

    # Summary covers every active entity plus a totals row, never inactive ones
    summary_entities = [row[0].value for row in wb["Summary"].iter_rows(min_row=2)]
    assert "Delta Development LLC" not in summary_entities
    assert summary_entities[-1] == "TOTAL"

    # Methodology lists the full spec — implemented AND pending — for honest coverage
    methodology_ids = [row[0].value for row in wb["Methodology"].iter_rows(min_row=2)]
    assert "T1-01" in methodology_ids
    assert "T1-15" in methodology_ids  # pending rule still listed
    statuses = {row[0].value: row[2].value for row in wb["Methodology"].iter_rows(min_row=2)}
    assert statuses["T1-15"] == "Pending data source"
    assert statuses["T1-01"] == "Implemented"
