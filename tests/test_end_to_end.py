"""End-to-end smoke test: drive the real CLI (skill.run.main) over a synthetic
export and confirm it ingests, runs the battery, and writes a workbook offline."""
import sys

import pandas as pd
from openpyxl import load_workbook

from skill.run import main


def _credit_memo_export() -> pd.DataFrame:
    # Matches the qb__credit_memos source mapping; $1,000 > credit_memo_threshold.
    return pd.DataFrame({
        "Trans #": ["90001"], "Date": ["2026-05-12"], "Name": ["Acme Lumber"],
        "Num": ["CM-1"], "Class": [""], "Account": ["2100 · Accounts Payable"],
        "Amount": [1000.00], "Memo": ["overbill credit"], "Last modified by": ["Megan"],
    })


def test_cli_runs_end_to_end(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    (data_dir / "hines-homes").mkdir(parents=True)            # a real registry entity id
    _credit_memo_export().to_csv(
        data_dir / "hines-homes" / "qb__credit_memos.csv", index=False)

    out = tmp_path / "exceptions.xlsx"
    monkeypatch.setattr(sys, "argv", [
        "skill.run", "--data-dir", str(data_dir), "--output", str(out),
        "--tier3", "off", "--store", "none", "--entity", "hines-homes"])

    main()   # the actual weekly-run entry point, offline (no AI, no Supabase)

    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary", "CRITICAL", "HIGH", "MEDIUM", "INFO",
                             "Methodology", "Run Info"]
    # The $1,000 credit memo surfaces as a T1-30 finding on the MEDIUM sheet.
    assert "T1-30" in [row[0].value for row in wb["MEDIUM"].iter_rows(min_row=2)]
    # Methodology proves Tier 1, 2, and 4 rules are all registered (honest coverage).
    rule_ids = {row[0].value for row in wb["Methodology"].iter_rows(min_row=2)}
    assert {"T1-30", "T2-02", "T2-10", "T4-02"} <= rule_ids
