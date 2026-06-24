"""Findings fingerprinting + disposition memory (suppress cleared / escalate recurrence)."""
import pytest
from openpyxl import load_workbook

from core.findings import Finding, Severity
from persistence import InMemoryFindingsStore, apply_disposition_memory
from persistence.findings_store import _pattern_key
from reporting.workbook import write_workbook
from rules.engine import run_all


@pytest.fixture
def findings(ctx):
    return run_all(ctx)


def find(findings, rule_id):
    return next(f for f in findings if f.rule_id == rule_id)


def entities_map(registry):
    return {e.id: e for e in registry}


# ---------------------------------------------------------------- fingerprint

def test_fingerprint_stable_and_transaction_sensitive():
    a = Finding("T1-01", Severity.CRITICAL, ["alpha"], "?", transactions=["TX-002", "TX-001"])
    b = Finding("T1-01", Severity.CRITICAL, ["alpha"], "different text",
                transactions=["TX-001", "TX-002"])
    assert a.fingerprint() == b.fingerprint()        # order-independent, text-independent
    c = Finding("T1-01", Severity.CRITICAL, ["alpha"], "?", transactions=["TX-003"])
    assert c.fingerprint() != a.fingerprint()


def test_fingerprint_uses_natural_key_when_no_transactions():
    a = Finding("T1-24", Severity.HIGH, ["alpha", "beta"], "?",
                details={"debtor": "alpha", "creditor": "beta"})
    b = Finding("T1-24", Severity.HIGH, ["beta", "alpha"], "?",
                details={"creditor": "beta", "debtor": "alpha"})
    assert a.fingerprint() == b.fingerprint()
    c = Finding("T1-24", Severity.HIGH, ["alpha", "beta"], "?",
                details={"debtor": "beta", "creditor": "alpha"})
    assert c.fingerprint() != a.fingerprint()


# ---------------------------------------------------------------- store round-trip

def test_in_memory_store_saves_new_open_findings(ctx):
    findings = run_all(ctx)
    store = InMemoryFindingsStore()
    store.save(findings)
    prior = store.load_prior()
    assert len(prior) == len(findings)
    assert set(prior["disposition"]) == {"open"}
    # idempotent: saving again adds nothing
    store.save(findings)
    assert len(store.load_prior()) == len(findings)


# ---------------------------------------------------------------- disposition memory

def test_no_prior_is_noop(ctx, registry, findings):
    kept, suppressed = apply_disposition_memory(findings, None, entities_map(registry))
    assert suppressed == []
    assert len(kept) == len(findings)


def test_exact_cleared_finding_is_suppressed(ctx, registry, findings):
    dup = find(findings, "T1-01")
    prior = InMemoryFindingsStore([{
        "fingerprint": dup.fingerprint(), "rule_id": "T1-01",
        "entity_ids": ["alpha"], "disposition": "legit",
        "details": {"vendor": "Acme Lumber"},
    }]).load_prior()

    kept, suppressed = apply_disposition_memory(findings, prior, entities_map(registry))
    assert dup.fingerprint() in {f.fingerprint() for f in suppressed}
    assert dup.fingerprint() not in {f.fingerprint() for f in kept}
    assert "previously reviewed" in suppressed[0].details["disposition_memory"]


def test_open_prior_is_not_suppressed(ctx, registry, findings):
    dup = find(findings, "T1-01")
    prior = InMemoryFindingsStore([{
        "fingerprint": dup.fingerprint(), "rule_id": "T1-01",
        "entity_ids": ["alpha"], "disposition": "open", "details": {},
    }]).load_prior()
    kept, suppressed = apply_disposition_memory(findings, prior, entities_map(registry))
    assert suppressed == []
    assert dup.fingerprint() in {f.fingerprint() for f in kept}


def test_recurrence_after_clear_escalates(ctx, registry, findings):
    off = find(findings, "T1-07")           # MEDIUM off-cycle payment, vendor present
    assert off.severity == Severity.MEDIUM
    vendor = off.details["vendor"]
    # A *different* finding (other fingerprint) of the same vendor pattern was cleared.
    prior = InMemoryFindingsStore([{
        "fingerprint": "previously-cleared-fp", "rule_id": "T1-07",
        "entity_ids": ["alpha"], "disposition": "error_corrected",
        "details": {"vendor": vendor},
    }]).load_prior()

    kept, suppressed = apply_disposition_memory(findings, prior, entities_map(registry))
    escalated = find(kept, "T1-07")
    assert escalated.severity == Severity.HIGH      # MEDIUM → HIGH
    assert "Recurs after a prior clear" in escalated.details["recurrence"]
    assert escalated not in suppressed


def test_vendorless_clear_does_not_escalate_unrelated(ctx, registry, findings):
    # A cleared finding with no vendor key must NOT escalate every later finding
    # of the same rule/entity — there's no discriminator to call it a recurrence.
    memo = find(findings, "T1-30")          # credit memo on beta, no vendor detail
    prior = InMemoryFindingsStore([{
        "fingerprint": "previously-cleared-fp", "rule_id": "T1-30",
        "entity_ids": ["beta"], "disposition": "error_corrected", "details": {},
    }]).load_prior()
    kept, _ = apply_disposition_memory(findings, prior, entities_map(registry))
    assert find(kept, "T1-30").severity == memo.severity        # unchanged
    assert "recurrence" not in find(kept, "T1-30").details


def test_pattern_key_requires_vendor():
    assert _pattern_key("T1-01", ["beta", "alpha"], {"vendor": "Acme"}) \
        == _pattern_key("T1-01", ["alpha", "beta"], {"vendor": "Acme"})
    assert _pattern_key("T1-01", ["alpha"], {"vendor": "Acme"}) \
        != _pattern_key("T1-01", ["alpha"], {"vendor": "Other"})
    assert _pattern_key("T1-30", ["beta"], {}) is None         # no vendor → no key


# ---------------------------------------------------------------- workbook surfacing

def test_workbook_lists_suppressed_findings(ctx, registry, tmp_path, findings):
    dup = find(findings, "T1-01")
    prior = InMemoryFindingsStore([{
        "fingerprint": dup.fingerprint(), "rule_id": "T1-01",
        "entity_ids": ["alpha"], "disposition": "legit", "details": {},
    }]).load_prior()
    kept, suppressed = apply_disposition_memory(findings, prior, entities_map(registry))

    path = write_workbook(kept, registry, tmp_path / "exc.xlsx", suppressed=suppressed)
    wb = load_workbook(path)
    assert "Dispositioned" in wb.sheetnames
    rule_ids = [row[0].value for row in wb["Dispositioned"].iter_rows(min_row=2)]
    assert "T1-01" in rule_ids
